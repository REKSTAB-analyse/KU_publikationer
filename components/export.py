import io
import csv
import pyarrow as pa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import streamlit as st

def build_table(rows: list, schema_fields: list) -> pa.Table:
    cols   = {name: [r.get(name) for r in rows] for name, _ in schema_fields}
    arrays = [pa.array(cols[name], type=dtype) for name, dtype in schema_fields]
    return pa.Table.from_arrays(arrays, names=[name for name, _ in schema_fields])
 
def rows_to_csv_bytes(rows: list, field_order: list) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=field_order, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r.get(k) for k in field_order})
    return buf.getvalue().encode("utf-8")
 
def rows_to_excel_bytes(rows: list, field_order: list, sheet_name: str = "Data") -> bytes:
    KU_RED = "901a1E"
 
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
 
    header_font  = Font(name="Arial", bold=False, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color=KU_RED, end_color=KU_RED)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_side    = Side(style="thin", color="DDDDDD")
    cell_border  = Border(bottom=thin_side)
    data_font    = Font(name="Arial", size=10)
    data_align   = Alignment(horizontal="left", vertical="center")
 
    for col_idx, col_name in enumerate(field_order, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
 
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(field_order, start=1):
            val        = row.get(col_name)
            cell       = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font  = data_font
            cell.alignment = data_align
            cell.border    = cell_border
            if isinstance(val, float):
                cell.number_format = "0.00"
            elif isinstance(val, int):
                cell.number_format = "0"
 
    for col_idx, col_name in enumerate(field_order, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(row.get(col_name, "") or "")) for row in rows),
        ) if rows else len(col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
 
    ws.freeze_panes = "A2"
 
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
 
def fmt_dk(value, decimals: int = 1) -> str:
    """Formater tal med dansk notation: punktum som tusindtalsseparator, komma som decimal."""
    if value is None:
        return ""
    formatted = f"{value:,.{decimals}f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")
 
def fmt_ui(value, decimals=1):
    """Formater tal til UI: punktum som decimal, ingen tusindseparator."""
    if value is None:
        return ""
    fmt = f"{{:.{decimals}f}}"
    return fmt.format(value)

def render_table_export(
    data: dict,
    row_label: str = "Enhed",
    col_order: list = None,
    col_labels: dict = None,
    filename: str = "data.xlsx",
    sheet_name: str = "Data",
    key: str = None,
) -> None:
    """
    Viser en "Se tabel"-expander med en tabel (kolonner i header, "row_label" som
    første kolonne) samt en "Download (.xlsx)"-knap.

    data: {row_navn: {kolonne: værdi}}, fx
          {"Forfatterpar (netværk)": {2021: 1863, 2022: 1900, ...}, ...}
          eller {"HUM": {"Open": 45.2, "Closed": 17.7, ...}, ...}
    col_order:  rækkefølge af kolonner. Default: sorteret automatisk.
    col_labels: {kolonne: visningsnavn}, fx OA_LABELS, hvis rå nøgler ikke skal
                bruges direkte som kolonneoverskrifter.
    key:        skal være unik pr. kald, hvis funktionen bruges flere gange på
                samme side (fx "Antal" og "Andel (%)" for samme chart).
    """
    if not data:
        return

    if col_order is None:
        cols = sorted({c for row in data.values() for c in row.keys()}, key=str)
    else:
        cols = col_order

    col_labels  = col_labels or {}
    header_cols = [col_labels.get(c, str(c)) for c in cols]
    field_order = [row_label] + header_cols

    rows = [
        {
            row_label: row_name,
            **{
                col_labels.get(c, str(c)): (row.get(c) if row.get(c) is not None else 0)
                for c in cols
            },
        }
        for row_name, row in data.items()
    ]

    with st.expander("Se tabel"):
        st.dataframe(rows, hide_index=True, width="stretch")

        excel_bytes = rows_to_excel_bytes(rows, field_order, sheet_name=sheet_name)
        st.download_button(
            "Download (.xlsx)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key or f"dl_{sheet_name}_{filename}",
        )
